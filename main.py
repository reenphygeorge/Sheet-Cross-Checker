from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Loading the 2 sheets from directory
wb1 = load_workbook('sheet1.xlsx')
ws1 = wb1.active
wb2 = load_workbook('sheet2.xlsx')
ws2 = wb2.active

# Global Variables
row1 = ""
row2 = ""   
contact1=""
contact2=""
name1=""
email1=""

# Function Definitions
def pos_detect():
    global row1, row2, contact1, contact2, name1, email1
    row1 = ws1.max_row
    row2 = ws2.max_row
    contact1 = 1;
    while (ws1[get_column_letter(contact1) + '1'].value).lower() != 'contact no':
        contact1 += 1

    contact2 = 1
    while (ws2[get_column_letter(contact2) + '1'].value).lower() != 'contact no':
        contact2 += 1

    name1 = 1
    while (ws1[get_column_letter(name1) + '1'].value).lower() != 'name':
        name1 += 1

    email1 = 1
    while (ws1[get_column_letter(email1) + '1'].value).lower() != 'email':
        email1 += 1

def check_writer():
    f = open('details.txt' , 'a')
    
    for i in range(2,row1+1):
        cn1 = ws1[get_column_letter(contact1) + str(i)].value
        for j in range(2,row2+1):
            cn2 = ws2[get_column_letter(contact2) + str(j)].value
            if cn1 == cn2:
                name = ws1[get_column_letter(name1) + str(i)].value
                email = ws1[get_column_letter(email1) + str(i)].value
                f.write("Name:"+name+"\nEmail:"+email+"\n\n")
                break;
    print("\n\nFile write successful")

# Invoke Functions
pos_detect();
check_writer();
